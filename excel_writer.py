import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import OUTPUT_EXCEL_PATH

COLUMNS = [
    "Enquiry", "Cust", "RFQ D", "Due D", "Due Time",
    "Delivery Time", "Status", "SN", "CPN", "Description",
    "Enq. Part No*", "Enq. Mfg", "Qty", "Unit",
    "ITEM/VALUE.(EVALUATION)", "TP", "Remark", "ACT.Due DT"
]

COL_WIDTHS  = [14, 10, 12, 12, 10, 14, 10, 5, 16, 40, 16, 28, 6, 6, 22, 8, 10, 12]
DATE_COLS   = {"RFQ D", "Due D", "ACT.Due DT"}
EXCEL_EPOCH = datetime(1899, 12, 30)

HDR_FILL  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN      = Side(style="thin", color="BBBBBB")
BDR       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FILL_W    = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
FILL_B    = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")
DATA_FONT = Font(name="Arial", size=10)
D_ALIGN   = Alignment(horizontal="left", vertical="center")


def _to_serial(d: datetime) -> int:
    return (d - EXCEL_EPOCH).days


def _write_header(ws):
    for ci, (col, w) in enumerate(zip(COLUMNS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = HDR_ALIGN
        cell.border    = BDR
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 28


def _write_row(ws, row_data: dict, row_num: int):
    fill = FILL_W if row_num % 2 == 0 else FILL_B
    for ci, col in enumerate(COLUMNS, 1):
        val  = row_data.get(col, "")
        cell = ws.cell(row=row_num, column=ci)
        if col in DATE_COLS:
            if isinstance(val, datetime):
                cell.value         = _to_serial(val)
                cell.number_format = "DD-MMM-YY"
            else:
                cell.value = str(val) if val else ""
        elif col == "SN":
            try:    cell.value = int(val)
            except: cell.value = val
        elif col == "Qty":
            try:
                f = float(val)
                cell.value = int(f) if f == int(f) else f
            except: cell.value = val
        else:
            cell.value = str(val) if val not in (None, "") else ""
        cell.font      = DATA_FONT
        cell.fill      = fill
        cell.alignment = D_ALIGN
        cell.border    = BDR
    ws.row_dimensions[row_num].height = 18


def append_rows(new_rows: list[dict]):
    """
    Appends new_rows to the Excel file at OUTPUT_EXCEL_PATH.
    Creates the file and writes the header if it does not exist yet.
    If the file exists, new rows are added after the last existing row.
    The file is saved immediately after writing.
    """
    os.makedirs(os.path.dirname(OUTPUT_EXCEL_PATH), exist_ok=True)

    if os.path.exists(OUTPUT_EXCEL_PATH):
        wb = openpyxl.load_workbook(OUTPUT_EXCEL_PATH)
        ws = wb.active
        next_row = ws.max_row + 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RFQ Log"
        ws.freeze_panes = "A2"
        _write_header(ws)
        next_row = 2

    for row_data in new_rows:
        _write_row(ws, row_data, next_row)
        next_row += 1

    try:
        wb.save(OUTPUT_EXCEL_PATH)
    except PermissionError:
        raise PermissionError(
            f"Cannot save to {OUTPUT_EXCEL_PATH} — the file is open in Excel. "
            f"Please close it and try again."
        )
    return OUTPUT_EXCEL_PATH

