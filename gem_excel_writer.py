import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import GEM_OUTPUT_EXCEL_PATH

# ── Column definitions ────────────────────────────────────────────────────────
COLUMNS = [
    "Enquiry", "Cust", "RFQ D", "Due D", "Time", "L/T", "STS", "SN",
    "CUST PART NO.", "Description", "Enq. Part No*", "Enq. Mfg",
    "Qty", "UOM", "TP", "Remarks",
]

COL_WIDTHS = [22, 18, 14, 14, 12, 6, 12, 5, 24, 45, 24, 12, 8, 12, 8, 55]

# Columns that should be centered
CENTER_COLS = {"SN", "Qty", "L/T", "TP", "Time", "UOM", "STS"}

# ── Styling constants ─────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN = Side(style="thin", color="BBBBBB")
BDR  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_W    = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
FILL_B    = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")
DATA_FONT = Font(name="Arial", size=9)

ALIGN_LEFT   = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")


def _write_header(ws):
    """Write the header row with styling and column widths."""
    for ci, (col, w) in enumerate(zip(COLUMNS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = HDR_ALIGN
        cell.border    = BDR
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 28


def _write_row(ws, row_data: dict, row_num: int):
    """Write a single data row with alternating row colors."""
    fill = FILL_W if row_num % 2 == 0 else FILL_B
    for ci, col in enumerate(COLUMNS, 1):
        val  = row_data.get(col, "")
        cell = ws.cell(row=row_num, column=ci)

        # Numeric handling for SN and Qty
        if col == "SN":
            try:
                cell.value = int(val)
            except (ValueError, TypeError):
                cell.value = val
        elif col == "Qty":
            try:
                f = float(val)
                cell.value = int(f) if f == int(f) else f
            except (ValueError, TypeError):
                cell.value = val
        else:
            cell.value = str(val) if val not in (None, "") else ""

        cell.font      = DATA_FONT
        cell.fill      = fill
        cell.alignment = ALIGN_CENTER if col in CENTER_COLS else ALIGN_LEFT
        cell.border    = BDR

    ws.row_dimensions[row_num].height = 22


def append_rows(new_rows: list[dict]) -> str:
    """
    Appends new_rows to the GeM Excel file at GEM_OUTPUT_EXCEL_PATH.
    Creates the file and writes the header if it does not exist yet.
    If the file exists, new rows are added after the last existing row.
    Returns the path to the saved file.
    """
    os.makedirs(os.path.dirname(GEM_OUTPUT_EXCEL_PATH), exist_ok=True)

    if os.path.exists(GEM_OUTPUT_EXCEL_PATH):
        wb = openpyxl.load_workbook(GEM_OUTPUT_EXCEL_PATH)
        ws = wb.active
        next_row = ws.max_row + 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "GeM Bid Extract"
        ws.freeze_panes = "A2"
        _write_header(ws)
        next_row = 2

    for row_data in new_rows:
        _write_row(ws, row_data, next_row)
        next_row += 1

    try:
        wb.save(GEM_OUTPUT_EXCEL_PATH)
    except PermissionError:
        raise PermissionError(
            f"Cannot save to {GEM_OUTPUT_EXCEL_PATH} — the file is open in Excel. "
            f"Please close it and try again."
        )
    return GEM_OUTPUT_EXCEL_PATH
